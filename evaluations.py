import os
import re
import pickle
from pathlib import Path

import numpy as np
import pandas as pd
import torch
from sklearn.linear_model import LogisticRegression, LogisticRegressionCV
from sklearn.model_selection import train_test_split
from sklearn import metrics

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from train_no_temp import CNTF


def logistic_regression(Wtrain, y_train, Wtest, y_test):
    classifier = LogisticRegressionCV(cv=5, Cs=10, class_weight='balanced')
    classifier.fit(Wtrain, y_train)
    pred_prob = classifier.predict_proba(Wtest)
    fpr, tpr, thresholds = metrics.roc_curve(y_test, pred_prob[:, 1], pos_label=1)
    auc = metrics.auc(fpr, tpr)
    return auc


def prediction_at_discharge(Ws, labels, classifier=logistic_regression):
    Ws_discharge = torch.cat([W.sum(dim=0).reshape(1, -1) for W in Ws]).numpy()
    best_auc = []
    for i in range(5):
        Wtrain, Wtest, labels_train, labels_test = train_test_split(Ws_discharge, labels, 
                                                                    train_size=0.8, test_size=0.2)
        best_auc.append(classifier(Wtrain, labels_train, Wtest, labels_test))
    return max(best_auc)


def load_combine_map_files(raw_file_dir, processed_file):
    lab_code2desc = pd.read_csv(os.path.join(raw_file_dir, 'labmap.csv'), index_col=0)
    rx_code2desc = pd.read_csv(os.path.join(raw_file_dir, 'rxmap.csv'), index_col=0)
    # dx_code2desc = pd.rad_csv(os.path.join(raw_file_dir, 'dxmap.csv'), index_col=0)
    dx_code2desc = pd.read_json(os.path.join(raw_file_dir, 'icd9.json')).set_index('code')
    with open(processed_file, 'rb') as f:
        infile = pickle.load(f)
        lab_idx2desc = {idx: lab_code2desc.loc[code]['label'] for code, idx in infile['labmap'].items()}
        rx_idx2code = {idx: rx_code2desc.loc[code]['drug_name'] for code, idx in infile['rxmap'].items()}
        dx_idx2code = {idx: dx_code2desc.loc[code]['title'] for code, idx in infile['dxmap'].items()}
    return lab_idx2desc, rx_idx2code, dx_idx2code


def interpret_phenotypes(*factors, item_idx2desc):
    phenotypes = []
    n_dims = len(factors)
    n_factors = factors[0].shape[1]
    item_sortidx = [np.argsort(-U, axis=0) for U in factors]
    # item_maps_inverse = [{idx: code} for items in item_maps for code, idx in items]

    for r in range(n_factors):
        phenotype_definition = []
        for j in range(n_dims):
            dim_j = []
            for idx in item_sortidx[j][:, r]:
                if factors[j][idx, r] > 1e-4:
                    dim_j.append((item_idx2desc[j][idx], factors[j][idx, r]))
                else:
                    break
            phenotype_definition.append(dim_j)
        phenotypes.append(phenotype_definition)
    return phenotypes


def coord(i, j):
    return '{}{}'.format(get_column_letter(j), i)


def phenotypes_to_excel_worksheet(phenotypes, dim_names, ws):
    # weights = list(map(lambda x:x[0], phenotypes))
    # for i, r in enumerate(np.argsort(weights)[::-1]):
    n_dims = len(phenotypes[0])
    for i, pheno_r in enumerate(phenotypes):
        ws.merge_cells(coord(1, (n_dims+1)*i+1)+':'+coord(1, (n_dims+1)*i+n_dims))
        ws[coord(1, (n_dims+1)*i+1)] = 'Phenotype {:d}'.format(i+1)
        ws[coord(1, (n_dims+1)*i+1)].font = Font(bold=True)
        ws[coord(1, (n_dims+1)*i+1)].alignment = Alignment(horizontal='center', vertical='center')



        for j, name in enumerate(dim_names):
            ws[coord(2, (n_dims+1)*i+j+1)] = name
            for k, (item, weight) in enumerate(pheno_r[j]):
                ws[coord(k+3, (n_dims+1)*i+j+1)] = '{}({:.3f})'.format(item, weight)
                ws.column_dimensions[get_column_letter((n_dims+1)*i+j+1)].width = 50


def phenotypes_to_excel_file(phenotypes, dim_names, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    phenotypes_to_excel_worksheet(phenotypes, dim_names, ws)
    wb.save(filepath)


if __name__ == '__main__':
    results_dir = Path('results/redo/notemp_R20')

    # determine data set used.
    # rawdata = './data/raw-180808'
    # if 'balanced' in results_dir and '2k' not in results_dir:
    #     processed_data = './data/processed-rx_lab_dx-balanced-180814.pkl'
    # elif '2k' in results_dir and 'balanced' not in results_dir:
    #     processed_data = './data/processed-rx_lab_dx-2k-180814.pkl'
    # else:
    #     raise RuntimeError('Data set not understood.')
    #
    # results = [f.name for f in os.scandir(results_dir) if f.is_dir()]
    #
    # def get_weighting_from_name(name):
    #     m = re.search('-alpha(.+?)_(.+?)-beta(.+?)-', name)
    #     if m is None:
    #         raise ValueError('Cannot extract weightings from the folder name.')
    #     return tuple(map(lambda x: float(m.group(x)), [1, 2, 3]))
    #
    # def get_rand_state_from_name(name):
    #     m = re.search('-rand(.+?)-', name)
    #     return int(m.group(1))


    lab_idx2desc, rx_idx2code, dx_idx2code = load_combine_map_files(r'D:\research_projects\cntf_phenotyping\data\raw-180808',
                                                                    r'D:\research_projects\cntf_phenotyping\data\processed-rx_lab_dx-balanced-180814.pkl')

    # result_dirs = [f.path for f in os.scandir(r'D:\temporalPhenotyping\results\temporal-50dimLSTM-balanced-b79062') if f.is_dir() ]
    result_dirs = [results_dir]
    # for result_dir in result_dirs:
        # infile = torch.load(os.path.join(result_dir, 'phenotypes.pt'))

        # Ul, Ud, Um = map(lambda x: x.numpy(), torch.load(os.path.join(result_dir, 'phenotypes.pt')))
        # Ul = Ul / np.linalg.norm(Ul, axis=0, ord=1)
        # Ul[Ul<0.01] = 0
        # Ud = Ud / np.linalg.norm(Ud, axis=0, ord=1)
        # Ud[Ud < 0.01] = 0
        # Um = Um / np.linalg.norm(Um, axis=0, ord=1)
        # Um[Um < 0.01] = 0
        # with open(os.path.join(result_dir, 'f.pkl'), 'rb') as f:
        #     infile = pickle.load(f)
        # Ud = infile['Ud'].numpy()
        # Um = infile['Um'].numpy()
        # Ul = infile['Ul'].numpy()

    model = torch.load(results_dir / 'final_model.pt')


    phentoypes = interpret_phenotypes(model.Ul.data.numpy(), model.Um.data.numpy(), item_idx2desc=[lab_idx2desc, rx_idx2code])
    phenotypes_to_excel_file(phentoypes, ['Labtests', 'Medications'], os.path.join(results_dir, 'phenotypes.xlsx'))
